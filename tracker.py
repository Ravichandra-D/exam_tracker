import os
import json
import time
import threading
import datetime
import zipfile
import smtplib
import tkinter as tk
from tkinter import messagebox
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

try:
    import pyautogui
    PYAUTOGUI_OK = True
except ImportError:
    PYAUTOGUI_OK = False

try:
    from pynput import keyboard, mouse
    PYNPUT_OK = True
except ImportError:
    PYNPUT_OK = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    import pystray
    from PIL import Image, ImageDraw
    TRAY_OK = True
except ImportError:
    TRAY_OK = False

# ── Email Config ──────────────────────────────────────────────────────────────
EMAIL_SENDER   = "ravichandra182001@gmail.com"
EMAIL_PASSWORD = "qqkkpchmorfxwdbd"
EMAIL_RECEIVER = "ravichandra182001@gmail.com"

# ── Suspicious Keys ───────────────────────────────────────────────────────────
SUSPICIOUS_COMBOS = {
    "Alt+Tab"    : "Switched applications",
    "Win Key"    : "Opened Start Menu",
    "Ctrl+C"     : "Copied content",
    "Ctrl+V"     : "Pasted content",
    "Alt+F4"     : "Attempted to close app",
    "Ctrl+Z"     : "Undo action",
}

SCREENSHOT_INTERVAL = 30

# ── Colors ────────────────────────────────────────────────────────────────────
BG_DARK     = "#0A0E1A"
BG_CARD     = "#111827"
BG_PANEL    = "#1A2235"
ACCENT_CYAN = "#00D4FF"
ACCENT_BLUE = "#3B82F6"
ACCENT_PURP = "#8B5CF6"
ACCENT_PINK = "#EC4899"
ACCENT_ORAN = "#F59E0B"
GREEN_OK    = "#10B981"
RED_STOP    = "#EF4444"
RED_ALERT   = "#FF2D2D"
TEXT_WHITE  = "#F1F5F9"
TEXT_GREY   = "#64748B"
TEXT_DIM    = "#334155"


class ExamTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mouse & Keyboard Event Tracker — Built by Ravichandra D")
        self.root.geometry("980x700")
        self.root.minsize(860, 620)
        self.root.configure(bg=BG_DARK)
        self.root.resizable(True, True)

        self.tracking         = False
        self.events           = []
        self.lock             = threading.Lock()
        self.session_start    = None
        self.log_file         = None
        self.screenshot_dir   = None
        self.log_dir          = None
        self.kb_listener      = None
        self.ms_listener      = None
        self.running          = False
        self.key_count        = 0
        self.click_count      = 0
        self.scroll_count     = 0
        self.screenshot_count = 0
        self.suspicious_count = 0
        self.elapsed          = 0
        self.keys_held        = set()
        self.tray_icon        = None
        self._last_suspicious = ""

        self._build_ui()
        self._animate_pulse()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.bind("<Unmap>", self._on_minimize)

    # ─────────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        # Header
        header = tk.Frame(self.root, bg=BG_DARK)
        header.pack(fill="x")

        bar = tk.Canvas(header, height=4, bg=BG_DARK, highlightthickness=0)
        bar.pack(fill="x")
        bar.bind("<Configure>", lambda e: self._draw_gradient_bar(bar))

        title_row = tk.Frame(header, bg=BG_DARK)
        title_row.pack(fill="x", padx=24, pady=(12, 8))

        self.pulse_canvas = tk.Canvas(title_row, width=14, height=14,
                                      bg=BG_DARK, highlightthickness=0)
        self.pulse_canvas.pack(side="left", padx=(0, 10))
        self.pulse_oval = self.pulse_canvas.create_oval(2, 2, 12, 12,
                                                        fill=TEXT_GREY, outline="")

        tk.Label(title_row, text="ExamGuard Pro",
                 font=("Courier New", 20, "bold"),
                 fg=ACCENT_CYAN, bg=BG_DARK).pack(side="left")
        tk.Label(title_row,
                 text="  Mouse & Keyboard Event Tracker — Built by Ravichandra D",
                 font=("Courier New", 10), fg=TEXT_GREY, bg=BG_DARK).pack(
                     side="left", pady=(4, 0))

        self.session_label = tk.Label(title_row, text="● IDLE",
                                      font=("Courier New", 10, "bold"),
                                      fg=TEXT_GREY, bg=BG_DARK)
        self.session_label.pack(side="right")

        tk.Frame(self.root, bg=TEXT_DIM, height=1).pack(fill="x")

        # Control bar
        ctrl_bar = tk.Frame(self.root, bg=BG_PANEL,
                            highlightbackground=TEXT_DIM, highlightthickness=1)
        ctrl_bar.pack(fill="x", padx=16, pady=(8, 8))

        btn_area = tk.Frame(ctrl_bar, bg=BG_PANEL)
        btn_area.pack(side="left", padx=12, pady=10)

        self.start_btn = self._make_button(btn_area, "▶  START",
                                           GREEN_OK, self._start_tracking)
        self.start_btn.pack(side="left", padx=(0, 6), ipadx=10)

        self.stop_btn = self._make_button(btn_area, "■  STOP",
                                          RED_STOP, self._stop_tracking,
                                          state="disabled")
        self.stop_btn.pack(side="left", padx=(0, 6), ipadx=10)

        self._make_button(btn_area, "🗑  CLEAR",
                          TEXT_GREY, self._clear_log).pack(
                              side="left", padx=(0, 6), ipadx=6)

        self._make_button(btn_area, "📊  REPORT",
                          ACCENT_BLUE, self._generate_excel).pack(
                              side="left", padx=(0, 6), ipadx=6)

        self._make_button(btn_area, "📧  SEND EMAIL",
                          ACCENT_PURP,
                          lambda: threading.Thread(
                              target=self._send_email, daemon=True).start()
                          ).pack(side="left", ipadx=6)

        cfg_area = tk.Frame(ctrl_bar, bg=BG_PANEL)
        cfg_area.pack(side="left", padx=16, pady=8)
        tk.Label(cfg_area, text="INTERVAL (sec):",
                 font=("Courier New", 8), fg=TEXT_GREY,
                 bg=BG_PANEL).pack(side="left")
        self.interval_var = tk.IntVar(value=SCREENSHOT_INTERVAL)
        tk.Spinbox(cfg_area, from_=10, to=300,
                   textvariable=self.interval_var,
                   width=4, font=("Courier New", 10),
                   bg=BG_CARD, fg=ACCENT_CYAN,
                   buttonbackground=BG_CARD,
                   relief="flat", bd=4).pack(side="left", padx=(4, 16))

        chk_area = tk.Frame(ctrl_bar, bg=BG_PANEL)
        chk_area.pack(side="left", pady=8)
        tk.Label(chk_area, text="TRACK:",
                 font=("Courier New", 8, "bold"),
                 fg=TEXT_GREY, bg=BG_PANEL).pack(side="left", padx=(0, 6))
        self.opt_keys   = self._chk(chk_area, "Keys")
        self.opt_clicks = self._chk(chk_area, "Clicks")
        self.opt_scroll = self._chk(chk_area, "Scrolls")
        self.opt_shots  = self._chk(chk_area, "Screenshots")

        info_area = tk.Frame(ctrl_bar, bg=BG_PANEL)
        info_area.pack(side="right", padx=12)
        self.info_file = tk.Label(info_area, text="Log: —",
                                  font=("Courier New", 7),
                                  fg=TEXT_GREY, bg=BG_PANEL)
        self.info_file.pack(anchor="e")
        self.info_dir = tk.Label(info_area, text="Shots: —",
                                 font=("Courier New", 7),
                                 fg=TEXT_GREY, bg=BG_PANEL)
        self.info_dir.pack(anchor="e")

        # Content
        content = tk.Frame(self.root, bg=BG_DARK)
        content.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # 5 stat cards
        stats_row = tk.Frame(content, bg=BG_DARK)
        stats_row.pack(fill="x", pady=(0, 10))
        self.stat_keys   = self._stat_card(stats_row, "KEYSTROKES",   "0", ACCENT_CYAN)
        self.stat_clicks = self._stat_card(stats_row, "CLICKS",        "0", ACCENT_BLUE)
        self.stat_scroll = self._stat_card(stats_row, "SCROLLS",       "0", ACCENT_PURP)
        self.stat_shots  = self._stat_card(stats_row, "SCREENSHOTS",   "0", ACCENT_ORAN)
        self.stat_alert  = self._stat_card(stats_row, "⚠ SUSPICIOUS",  "0", RED_ALERT)

        mid_row = tk.Frame(content, bg=BG_DARK)
        mid_row.pack(fill="both", expand=True)

        # Left panel
        left_panel = tk.Frame(mid_row, bg=BG_DARK, width=270)
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        left_panel.pack_propagate(False)

        timer_card = tk.Frame(left_panel, bg=BG_CARD,
                              highlightbackground=TEXT_DIM, highlightthickness=1)
        timer_card.pack(fill="x", pady=(0, 8))

        tk.Label(timer_card, text="SESSION DURATION",
                 font=("Courier New", 8, "bold"),
                 fg=TEXT_GREY, bg=BG_CARD).pack(anchor="w", padx=14, pady=(12, 2))
        
        # --- FIX: Downscale time text font from 30 to 24 so text scales completely inside boundaries ---
        self.timer_label = tk.Label(timer_card, text="00:00:00",
                                    font=("Courier New", 24, "bold"),
                                    fg=ACCENT_CYAN, bg=BG_CARD)
        self.timer_label.pack(padx=14, pady=(2, 10))
        tk.Label(timer_card, text="NEXT SCREENSHOT IN",
                 font=("Courier New", 8),
                 fg=TEXT_GREY, bg=BG_CARD).pack(anchor="w", padx=14)

        pb_frame = tk.Frame(timer_card, bg=BG_PANEL, height=10,
                            highlightbackground=TEXT_DIM, highlightthickness=1)
        pb_frame.pack(fill="x", padx=14, pady=(4, 12))
        pb_frame.pack_propagate(False)
        self.pb_fill = tk.Frame(pb_frame, bg=ACCENT_ORAN, height=10)
        self.pb_fill.place(x=0, y=0, relwidth=1.0, height=10)

        # Suspicious panel
        sus_card = tk.Frame(left_panel, bg=BG_CARD,
                            highlightbackground=RED_ALERT, highlightthickness=1)
        sus_card.pack(fill="both", expand=True)

        sus_hdr = tk.Frame(sus_card, bg="#1a0808")
        sus_hdr.pack(fill="x")
        tk.Label(sus_hdr, text="⚠  SUSPICIOUS ACTIVITY",
                 font=("Courier New", 8, "bold"),
                 fg=RED_ALERT, bg="#1a0808").pack(anchor="w", padx=12, pady=6)

        self.sus_box = tk.Text(sus_card, bg=BG_CARD, fg=RED_ALERT,
                               font=("Courier New", 8),
                               state="disabled", relief="flat", bd=0,
                               padx=8, pady=6, wrap="word")
        self.sus_box.pack(fill="both", expand=True)

        # Log panel
        log_col = tk.Frame(mid_row, bg=BG_DARK)
        log_col.pack(side="left", fill="both", expand=True)

        log_hdr = tk.Frame(log_col, bg=BG_DARK)
        log_hdr.pack(fill="x", pady=(0, 4))
        tk.Label(log_hdr, text="LIVE EVENT LOG",
                 font=("Courier New", 9, "bold"),
                 fg=TEXT_GREY, bg=BG_DARK).pack(side="left")
        tk.Label(log_hdr, text="  real-time feed",
                 font=("Courier New", 8),
                 fg=TEXT_DIM, bg=BG_DARK).pack(side="left")
        self.email_status = tk.Label(log_hdr, text="",
                                     font=("Courier New", 8),
                                     fg=GREEN_OK, bg=BG_DARK)
        self.email_status.pack(side="right")

        log_frame = tk.Frame(log_col, bg=BG_CARD,
                             highlightbackground=TEXT_DIM, highlightthickness=1)
        log_frame.pack(fill="both", expand=True)

        self.log_box = tk.Text(log_frame, bg=BG_CARD, fg=TEXT_WHITE,
                               font=("Courier New", 9),
                               state="disabled", relief="flat", bd=0,
                               selectbackground=ACCENT_BLUE,
                               wrap="word", padx=10, pady=8)
        self.log_box.pack(fill="both", expand=True)

        self.log_box.tag_config("key",    foreground=ACCENT_CYAN)
        self.log_box.tag_config("click",  foreground=ACCENT_BLUE)
        self.log_box.tag_config("scroll", foreground=ACCENT_PURP)
        self.log_box.tag_config("shot",   foreground=ACCENT_ORAN)
        self.log_box.tag_config("sys",    foreground=GREEN_OK)
        self.log_box.tag_config("time",   foreground=TEXT_GREY)
        self.log_box.tag_config("alert",  foreground=RED_ALERT)

        # Status bar
        status_bar = tk.Frame(self.root, bg=BG_PANEL, height=28)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        self.status_label = tk.Label(status_bar,
                                     text="  Ready — Click START to begin",
                                     font=("Courier New", 8),
                                     fg=TEXT_GREY, bg=BG_PANEL, anchor="w")
        self.status_label.pack(side="left", fill="y")
        tk.Label(status_bar, text="ExamGuard Pro v2.0  ",
                 font=("Courier New", 8),
                 fg=TEXT_DIM, bg=BG_PANEL).pack(side="right")

    # ── Widget helpers ────────────────────────────────────────────────────────
    def _stat_card(self, parent, label, value, color):
        card = tk.Frame(parent, bg=BG_CARD,
                        highlightbackground=color, highlightthickness=1)
        card.pack(side="left", expand=True, fill="both", padx=(0, 6))
        tk.Canvas(card, height=3, bg=color,
                  highlightthickness=0).pack(fill="x")
        val = tk.Label(card, text=value,
                       font=("Courier New", 22, "bold"),
                       fg=color, bg=BG_CARD)
        val.pack(pady=(6, 0))
        tk.Label(card, text=label,
                 font=("Courier New", 7, "bold"),
                 fg=TEXT_GREY, bg=BG_CARD).pack(pady=(0, 8))
        return val

    def _make_button(self, parent, text, color, command, state="normal"):
        btn = tk.Button(parent, text=text,
                        font=("Courier New", 9, "bold"),
                        fg=BG_DARK, bg=color,
                        activebackground=color, activeforeground=BG_DARK,
                        relief="flat", bd=0, cursor="hand2",
                        pady=7, state=state, command=command)
        btn.bind("<Enter>", lambda e: btn.configure(
            bg=self._lighten(color)) if str(btn["state"]) != "disabled" else None)
        btn.bind("<Leave>", lambda e: btn.configure(bg=color))
        return btn

    def _chk(self, parent, text):
        var = tk.BooleanVar(value=True)
        tk.Checkbutton(parent, text=text, variable=var,
                       font=("Courier New", 8),
                       fg=TEXT_WHITE, bg=BG_PANEL,
                       activebackground=BG_PANEL,
                       activeforeground=ACCENT_CYAN,
                       selectcolor=BG_CARD,
                       relief="flat").pack(side="left", padx=(0, 10))
        return var

    def _lighten(self, hex_color):
        try:
            r = min(255, int(hex_color[1:3], 16) + 30)
            g = min(255, int(hex_color[3:5], 16) + 30)
            b = min(255, int(hex_color[5:7], 16) + 30)
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return hex_color

    def _draw_gradient_bar(self, canvas):
        canvas.delete("all")
        w = canvas.winfo_width()
        colors = [ACCENT_CYAN, ACCENT_BLUE, ACCENT_PURP, ACCENT_PINK]
        seg = max(1, w // len(colors))
        for i, c in enumerate(colors):
            canvas.create_rectangle(i * seg, 0, (i+1) * seg, 4,
                                    fill=c, outline="")

    # ── Pulse animation ───────────────────────────────────────────────────────
    def _animate_pulse(self):
        self.pulse_canvas.itemconfig(
            self.pulse_oval,
            fill=GREEN_OK if self.tracking else TEXT_GREY)
        self.root.after(800, self._animate_pulse2)

    def _animate_pulse2(self):
        self.pulse_canvas.itemconfig(
            self.pulse_oval,
            fill=BG_DARK if self.tracking else TEXT_DIM)
        self.root.after(800, self._animate_pulse)

    # ── Minimize to tray ──────────────────────────────────────────────────────
    def _on_minimize(self, event):
        if self.root.state() == "iconic" and TRAY_OK:
            self.root.withdraw()
            self._show_tray()

    def _show_tray(self):
        if self.tray_icon:
            return
        try:
            img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            color = (16, 185, 129) if self.tracking else (100, 116, 139)
            draw.ellipse([8, 8, 56, 56], fill=color)

            menu = pystray.Menu(
                pystray.MenuItem("Open ExamGuard Pro",
                                 self._restore_from_tray, default=True),
                pystray.MenuItem("Stop & Exit", self._quit_from_tray)
            )
            self.tray_icon = pystray.Icon(
                "ExamGuard", img, "ExamGuard Pro — Running", menu)
            threading.Thread(target=self.tray_icon.run, daemon=True).start()
        except Exception:
            self.root.deiconify()

    def _restore_from_tray(self, icon=None, item=None):
        if self.tray_icon:
            self.tray_icon.stop()
            self.tray_icon = None
        self.root.after(0, self.root.deiconify)

    def _quit_from_tray(self, icon=None, item=None):
        if self.tray_icon:
            self.tray_icon.stop()
        self.root.after(0, self._on_close)

    # ── Logging ───────────────────────────────────────────────────────────────
    def _append_log(self, text, tag):
        def _do():
            self.log_box.configure(state="normal")
            ts = datetime.datetime.now().strftime("%H:%M:%S")
            self.log_box.insert("end", f"[{ts}] ", "time")
            self.log_box.insert("end", text + "\n", tag)
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.root.after(0, _do)

    def _append_suspicious_panel(self, text):
        def _do():
            self.sus_box.configure(state="normal")
            ts = datetime.datetime.now().strftime("%H:%M:%S")
            self.sus_box.insert("end", f"[{ts}] {text}\n")
            self.sus_box.see("end")
            self.sus_box.configure(state="disabled")
        self.root.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _save_log(self):
        if self.log_file:
            with self.lock:
                with open(self.log_file, "w") as f:
                    json.dump(self.events, f, indent=2)

    def _log_event(self, event_type, data):
        entry = {
            "timestamp": datetime.datetime.now().isoformat(),
            "type": event_type,
            "data": data
        }
        with self.lock:
            self.events.append(entry)
        self._save_log()

    # ── Suspicious detection ──────────────────────────────────────────────────
    def _flag_suspicious(self, combo, reason):
        key = combo + reason
        if key == self._last_suspicious:
            return
        self._last_suspicious = key
        self.root.after(3000, lambda: setattr(self, "_last_suspicious", ""))

        self.suspicious_count += 1
        self.root.after(0, lambda: self.stat_alert.configure(text=str(self.suspicious_count)))
        self._log_event("suspicious", {"combo": combo, "reason": reason})
        
        self._append_log(f"⚠ SUSPICIOUS: {combo} — {reason}", "alert")
        self._append_suspicious_panel(f"{combo} — {reason}")
        
        def show_popup():
            messagebox.showwarning(
                "⚠ Suspicious Activity Detected",
                f"Activity Detected: {combo}\nReason: {reason}\n\nThis event has been logged."
            )
        threading.Thread(target=show_popup, daemon=True).start()

    # ── Keyboard ─────────────────────────────────────────────────────────────
    def _on_key_press(self, key):
        try:
            self.keys_held.add(key.char)
        except AttributeError:
            self.keys_held.add(key)

        held = self.keys_held
        K = keyboard.Key

        # Combo detection modifiers
        alt_held  = K.alt in held or K.alt_l in held or K.alt_r in held
        ctrl_held = K.ctrl in held or K.ctrl_l in held or K.ctrl_r in held
        win_held  = K.cmd in held or getattr(K, "cmd_l", None) in held

        # --- FIX: Account for ASCII control characters when Ctrl is held ---
        c_pressed = "c" in held or "\x03" in held
        v_pressed = "v" in held or "\x16" in held
        z_pressed = "z" in held or "\x1a" in held

        if alt_held and K.tab in held:
            self._flag_suspicious("Alt+Tab", SUSPICIOUS_COMBOS["Alt+Tab"])
        elif alt_held and K.f4 in held:
            self._flag_suspicious("Alt+F4", SUSPICIOUS_COMBOS["Alt+F4"])
        elif ctrl_held and c_pressed:
            self._flag_suspicious("Ctrl+C", SUSPICIOUS_COMBOS["Ctrl+C"])
        elif ctrl_held and v_pressed:
            self._flag_suspicious("Ctrl+V", SUSPICIOUS_COMBOS["Ctrl+V"])
        elif ctrl_held and z_pressed:
            self._flag_suspicious("Ctrl+Z", SUSPICIOUS_COMBOS["Ctrl+Z"])
        elif win_held:
            self._flag_suspicious("Win Key", SUSPICIOUS_COMBOS["Win Key"])

        if not self.opt_keys.get():
            return

        try:
            # --- FIX: Clean up display string for control characters ---
            if key.char == "\x03":
                k_str = "Ctrl+C"
            elif key.char == "\x16":
                k_str = "Ctrl+V"
            elif key.char == "\x1a":
                k_str = "Ctrl+Z"
            elif key.char and ord(key.char) < 32:
                k_str = f"Ctrl+{chr(ord(key.char) + 96).upper()}"
            else:
                k_str = key.char
        except AttributeError:
            k_str = str(key)

        if k_str is None:
            return

        self.key_count += 1
        self.root.after(0, lambda: self.stat_keys.configure(text=str(self.key_count)))
        self._log_event("key_press", {"key": k_str})
        self._append_log(f"KEY  → {k_str}", "key")

    def _on_key_release(self, key):
        try:
            # --- FIX: Ensure virtual character cleanups reset held storage arrays completely ---
            if key.char == "\x03":
                self.keys_held.discard("\x03")
                self.keys_held.discard("c")
            elif key.char == "\x16":
                self.keys_held.discard("\x16")
                self.keys_held.discard("v")
            elif key.char == "\x1a":
                self.keys_held.discard("\x1a")
                self.keys_held.discard("z")
            else:
                self.keys_held.discard(key.char)
        except AttributeError:
            self.keys_held.discard(key)

        # Safety flush: If the actual physical Ctrl keys are released, flush them all
        if key in (keyboard.Key.ctrl, keyboard.Key.ctrl_l, keyboard.Key.ctrl_r):
            self.keys_held.discard(keyboard.Key.ctrl)
            self.keys_held.discard(keyboard.Key.ctrl_l)
            self.keys_held.discard(keyboard.Key.ctrl_r)
            self.keys_held.discard("\x03")
            self.keys_held.discard("\x16")
            self.keys_held.discard("\x1a")

    # ── Mouse ─────────────────────────────────────────────────────────────────
    def _on_click(self, x, y, button, pressed):
        if not self.opt_clicks.get() or not pressed:
            return
        self.click_count += 1
        self.root.after(0, lambda: self.stat_clicks.configure(
            text=str(self.click_count)))
        self._log_event("mouse_click", {"x": x, "y": y, "button": str(button)})
        self._append_log(f"CLICK → ({x}, {y})  {button}", "click")

    def _on_scroll(self, x, y, dx, dy):
        if not self.opt_scroll.get():
            return
        self.scroll_count += 1
        self.root.after(0, lambda: self.stat_scroll.configure(
            text=str(self.scroll_count)))
        direction = "↑ up" if dy > 0 else "↓ down"
        self._log_event("mouse_scroll", {"x": x, "y": y, "direction": direction})
        self._append_log(f"SCROLL {direction} at ({x}, {y})", "scroll")

    # ── Screenshot thread ─────────────────────────────────────────────────────
    def _screenshot_loop(self):
        interval = self.interval_var.get()
        elapsed_in_interval = 0
        while self.running:
            time.sleep(1)
            elapsed_in_interval += 1
            pct = elapsed_in_interval / interval
            self.root.after(0, lambda p=pct: self.pb_fill.place(
                x=0, y=0, relwidth=min(p, 1.0), height=10))
            if elapsed_in_interval >= interval:
                elapsed_in_interval = 0
                if self.opt_shots.get() and PYAUTOGUI_OK:
                    ts = datetime.datetime.now().strftime("%H-%M-%S")
                    self.screenshot_count += 1
                    fname = f"screenshot_{self.screenshot_count:03d}_{ts}.png"
                    path = os.path.join(self.screenshot_dir, fname)
                    pyautogui.screenshot(path)
                    self.root.after(0, lambda: self.stat_shots.configure(
                        text=str(self.screenshot_count)))
                    self._log_event("screenshot", {"file": fname})
                    self._append_log(f"📸 SCREENSHOT → {fname}", "shot")

    # ── Timer thread ──────────────────────────────────────────────────────────
    def _timer_loop(self):
        while self.running:
            time.sleep(1)
            self.elapsed += 1
            h = self.elapsed // 3600
            m = (self.elapsed % 3600) // 60
            s = self.elapsed % 60
            self.root.after(0, lambda t=f"{h:02d}:{m:02d}:{s:02d}":
                            self.timer_label.configure(text=t))

    # ── Excel report ──────────────────────────────────────────────────────────
    def _generate_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Missing Library",
                                 "Run: pip install openpyxl")
            return None
        if not self.events:
            messagebox.showinfo("No Data", "No events recorded yet.")
            return None

        out_dir = self.log_dir or os.path.expanduser("~")
        path = os.path.join(out_dir, f"report_{self.session_start}.xlsx")
        wb = openpyxl.Workbook()

        # Sheet 1 — All Events
        ws1 = wb.active
        ws1.title = "All Events"
        ws1.column_dimensions["A"].width = 26
        ws1.column_dimensions["B"].width = 18
        ws1.column_dimensions["C"].width = 44

        hfill = PatternFill("solid", fgColor="0A0E1A")
        hfont = Font(bold=True, color="00D4FF", name="Courier New")
        for col, title in enumerate(["TIMESTAMP", "EVENT TYPE", "DETAILS"], 1):
            c = ws1.cell(row=1, column=col, value=title)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center")

        color_map = {
            "key_press":    ("00D4FF", "0d1a2e"),
            "mouse_click":  ("3B82F6", "0d1828"),
            "mouse_scroll": ("8B5CF6", "110d28"),
            "screenshot":   ("F59E0B", "1e1a0d"),
            "suspicious":   ("FF2D2D", "2a0808"),
            "session_start":("10B981", "0d1a10"),
            "session_end":  ("10B981", "0d1a10"),
        }

        for r, ev in enumerate(self.events, 2):
            ts    = ev.get("timestamp", "")
            etype = ev.get("type", "")
            data  = ev.get("data", {})
            detail = ", ".join(f"{k}={v}" for k, v in data.items())
            fg, bg = color_map.get(etype, ("F1F5F9", "111827"))
            fill  = PatternFill("solid", fgColor=bg)
            font  = Font(color=fg, name="Courier New", size=9)
            for col, val in enumerate([ts, etype, detail], 1):
                c = ws1.cell(row=r, column=col, value=val)
                c.font  = font
                c.fill  = fill
                c.alignment = Alignment(horizontal="left")

        # Sheet 2 — Summary
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 22

        rows = [
            ("SESSION SUMMARY", "", True),
            ("Session Start",   self.session_start or "—", False),
            ("Session End",     datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), False),
            ("Duration",        f"{self.elapsed // 60}m {self.elapsed % 60}s", False),
            ("", "", False),
            ("ACTIVITY COUNTS", "", True),
            ("Total Keystrokes",   self.key_count,        False),
            ("Total Mouse Clicks", self.click_count,      False),
            ("Total Scrolls",      self.scroll_count,     False),
            ("Total Screenshots",  self.screenshot_count, False),
            ("Suspicious Events",  self.suspicious_count, False),
            ("Total Events",       len(self.events),      False),
        ]
        for r, (label, value, is_header) in enumerate(rows, 1):
            c1 = ws2.cell(row=r, column=1, value=label)
            c2 = ws2.cell(row=r, column=2, value=value)
            if is_header:
                c1.font = Font(bold=True, color="00D4FF",
                               name="Courier New", size=11)
                for c in (c1, c2):
                    c.fill = PatternFill("solid", fgColor="0A0E1A")
            else:
                c1.font = Font(color="64748B", name="Courier New", size=9)
                c2.font = Font(color="F1F5F9", name="Courier New",
                               size=9, bold=True)
                for c in (c1, c2):
                    c.fill = PatternFill("solid", fgColor="111827")

        wb.save(path)
        self._append_log(f"📊 Excel report → {os.path.basename(path)}", "sys")
        return path

    # ── Zip ───────────────────────────────────────────────────────────────────
    def _create_zip(self):
        if not self.log_dir:
            return None
        zip_path = os.path.join(
            self.log_dir, f"session_report_{self.session_start}.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            if self.log_file and os.path.exists(self.log_file):
                zf.write(self.log_file, os.path.basename(self.log_file))
            excel = os.path.join(self.log_dir,
                                 f"report_{self.session_start}.xlsx")
            if os.path.exists(excel):
                zf.write(excel, os.path.basename(excel))
            if self.screenshot_dir and os.path.exists(self.screenshot_dir):
                for f in os.listdir(self.screenshot_dir):
                    zf.write(os.path.join(self.screenshot_dir, f),
                             os.path.join("screenshots", f))
        return zip_path

    # ── Send email ────────────────────────────────────────────────────────────
    def _send_email(self):
        self.root.after(0, lambda: self.email_status.configure(
            text="📧 Sending...", fg=ACCENT_ORAN))
        try:
            self._generate_excel()
            zip_path = self._create_zip()

            msg = MIMEMultipart()
            msg["From"]    = EMAIL_SENDER
            msg["To"]      = EMAIL_RECEIVER
            msg["Subject"] = (f"ExamGuard Pro — Session Report "
                              f"{self.session_start}")

            body = f"""
ExamGuard Pro — Session Report
================================
Built by: Ravichandra D

Session Start   : {self.session_start}
Duration        : {self.elapsed // 60}m {self.elapsed % 60}s

ACTIVITY SUMMARY
----------------
Keystrokes      : {self.key_count}
Mouse Clicks    : {self.click_count}
Scrolls         : {self.scroll_count}
Screenshots     : {self.screenshot_count}
Suspicious Acts : {self.suspicious_count}
Total Events    : {len(self.events)}

Attached: JSON log + Excel report + Screenshots (zipped)

— ExamGuard Pro v2.0
"""
            msg.attach(MIMEText(body, "plain"))

            if zip_path and os.path.exists(zip_path):
                with open(zip_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                                f"attachment; filename="
                                f"{os.path.basename(zip_path)}")
                msg.attach(part)

            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(EMAIL_SENDER, EMAIL_PASSWORD)
                server.sendmail(EMAIL_SENDER, EMAIL_RECEIVER, msg.as_string())

            self.root.after(0, lambda: self.email_status.configure(
                text="✅ Email sent!", fg=GREEN_OK))
            self._append_log("📧 Report emailed to ravichandra182001@gmail.com", "sys")
            self.status_label.configure(
                text="  ✅ Session report sent to your email!")

        except Exception as e:
            self.root.after(0, lambda: self.email_status.configure(
                text="❌ Email failed", fg=RED_STOP))
            self._append_log(f"❌ Email error: {e}", "alert")
            self.root.after(0, lambda err=str(e): messagebox.showerror(
                "Email Error", f"Failed to send:\n{err}"))

    # ── Start ─────────────────────────────────────────────────────────────────
    def _start_tracking(self):
        if self.tracking:
            return
        self.tracking = True
        self.running  = True
        self.elapsed  = 0
        self.key_count = self.click_count = self.scroll_count = 0
        self.screenshot_count = self.suspicious_count = 0
        self.events    = []
        self.keys_held = set()
        self._last_suspicious = ""

        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.session_start = ts
        self.log_dir = os.path.join(
            os.path.expanduser("~"), "Desktop", "exam_logs")
        os.makedirs(self.log_dir, exist_ok=True)
        self.log_file = os.path.join(self.log_dir, f"session_{ts}.json")
        self.screenshot_dir = os.path.join(
            self.log_dir, f"screenshots_{ts}")
        os.makedirs(self.screenshot_dir, exist_ok=True)

        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.session_label.configure(text="● RECORDING", fg=GREEN_OK)
        self.status_label.configure(
            text="  Tracking active — all events being logged")
        self.timer_label.configure(text="00:00:00")
        for s in (self.stat_keys, self.stat_clicks, self.stat_scroll,
                  self.stat_shots, self.stat_alert):
            s.configure(text="0")
        self.info_file.configure(text=f"Log: session_{ts[:10]}.json")
        self.info_dir.configure(text=f"Shots: screenshots_{ts[:10]}")
        self.email_status.configure(text="")

        self._log_event("session_start", {"time": ts})
        self._append_log(f"Session started — {ts}", "sys")

        threading.Thread(target=self._timer_loop, daemon=True).start()
        threading.Thread(target=self._screenshot_loop, daemon=True).start()

        if PYNPUT_OK:
            self.ms_listener = mouse.Listener(
                on_click=self._on_click, on_scroll=self._on_scroll)
            self.ms_listener.start()
            self.kb_listener = keyboard.Listener(
                on_press=self._on_key_press,
                on_release=self._on_key_release)
            self.kb_listener.start()

    # ── Stop ──────────────────────────────────────────────────────────────────
    def _stop_tracking(self):
        if not self.tracking:
            return
        self.tracking = False
        self.running  = False

        if self.kb_listener:
            self.kb_listener.stop()
        if self.ms_listener:
            self.ms_listener.stop()

        self._log_event("session_end", {
            "time":              datetime.datetime.now().isoformat(),
            "total_keystrokes":  self.key_count,
            "total_clicks":      self.click_count,
            "total_scrolls":     self.scroll_count,
            "total_screenshots": self.screenshot_count,
            "suspicious_events": self.suspicious_count,
        })
        self._save_log()

        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.session_label.configure(text="● STOPPED", fg=RED_STOP)
        self.pb_fill.place(x=0, y=0, relwidth=0, height=10)
        self._append_log(
            f"Session ended — {self.key_count} keys, "
            f"{self.click_count} clicks, "
            f"{self.suspicious_count} suspicious events", "sys")
        self.status_label.configure(
            text="  Session saved. Preparing email report...")

        threading.Thread(target=self._send_email, daemon=True).start()

    def _on_close(self):
        if self.tracking:
            self._stop_tracking()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = ExamTrackerApp(root)
    root.mainloop()